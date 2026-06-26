from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="217346")  # Excel green
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_MONO = Font(name="Courier New")
_CENTER = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def build_excel(code_list: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Codigos Escaneados"

    headers = ["#", "Código", "Tipo", "Fecha_Hora", "Sesión"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    ws.freeze_panes = "A2"

    for item in code_list:
        ws.append([
            item["order"],
            item["code"],
            item["type"],
            item["timestamp"],
            item["session"],
        ])

    # Monospace for code column
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.font = _MONO

    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    cell = ws.cell(row=summary_row, column=1, value=f"Total únicos: {len(code_list)}")
    cell.font = _BOLD

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
